"""
Train the XGBoost model and save it to backend/model.joblib.

Run from the project root:
    python backend/train_model.py

Requires: data/Raw Data.xlsx
Produces: backend/model.joblib
"""
import warnings
warnings.filterwarnings("ignore")

from pathlib import Path

import joblib
import numpy as np
import openpyxl
import pandas as pd
from imblearn.over_sampling import SMOTE
from sklearn.metrics import accuracy_score, f1_score, roc_auc_score
from sklearn.model_selection import train_test_split
from xgboost import XGBClassifier

PROJECT_ROOT = Path(__file__).parent.parent
DATA_PATH    = PROJECT_ROOT / "data" / "Raw Data.xlsx"
MODEL_OUT    = Path(__file__).parent / "model.joblib"

FEATURE_COLS = [
    "Q1","Q2","Q3","Q4","Q5","Q6","Q7","Q8","Q9","Q10",
    "sas_total","usage_duration","social_media_usage",
    "frequent_access","age","gender_encoded",
]


def load_data(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True)

    ws1   = wb["Sociodemographic Data"]
    rows1 = list(ws1.iter_rows(values_only=True))
    df_socio = (
        pd.DataFrame([r for r in rows1[2:] if r[0] is not None], columns=rows1[1])
        .rename(columns={
            "No.":                   "ID",
            "Smartphone-usage Duration": "usage_duration",
            "Social Media Usage":    "social_media_usage",
            "The Most Frequent Access": "frequent_access",
            "Gender":                "gender_socio",
            "Age":                   "age",
            "Class":                 "grade",
        })
        .drop(columns=["Name/Initial"])
    )

    ws2   = wb["Smartphone Addiction"]
    rows2 = list(ws2.iter_rows(values_only=True))
    data2 = []
    for r in rows2[3:]:
        if r[0] is not None and isinstance(r[0], int):
            qs = [r[i] for i in range(2, 12)]
            data2.append({
                "ID": r[0], "gender": r[1],
                **{f"Q{i+1}": qs[i] for i in range(10)},
                "sas_total": sum(x for x in qs if isinstance(x, (int, float))),
                "addicted":  r[13],
            })
    df_add = pd.DataFrame(data2)

    df = pd.merge(df_socio, df_add, on="ID")
    df["gender_encoded"] = (df["gender"] == "P").astype(int)
    return df


def train_and_save() -> None:
    print(f"Loading data from {DATA_PATH} ...")
    df = load_data(DATA_PATH)
    print(f"Dataset: {df.shape[0]} records")

    X = df[FEATURE_COLS].copy()
    y = df["addicted"].copy()

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.10, random_state=42, stratify=y
    )

    X_train_bal, y_train_bal = SMOTE(random_state=42).fit_resample(X_train, y_train)
    print(f"After SMOTE: {pd.Series(y_train_bal).value_counts().to_dict()}")

    model = XGBClassifier(
        n_estimators=100, max_depth=4, learning_rate=0.1,
        subsample=0.8, colsample_bytree=0.8,
        eval_metric="logloss", random_state=42, verbosity=0,
    )
    model.fit(X_train_bal, y_train_bal)

    y_pred = model.predict(X_test)
    y_prob = model.predict_proba(X_test)[:, 1]
    print(f"Test Accuracy : {accuracy_score(y_test, y_pred):.4f}")
    print(f"Test F1       : {f1_score(y_test, y_pred):.4f}")
    print(f"Test AUC-ROC  : {roc_auc_score(y_test, y_prob):.4f}")

    joblib.dump(model, MODEL_OUT)
    print(f"\nModel saved → {MODEL_OUT}")


if __name__ == "__main__":
    train_and_save()
